from PyPtt import PTT, NewIndex
from datetime import datetime, timedelta, date
import openpyxl
import os
from dataclasses import dataclass
from typing import Dict, Set, Optional, Tuple
import logging
from pathlib import Path
import time
from tqdm import tqdm

@dataclass
class CrawlerConfig:
    """爬蟲設定"""
    ptt_id: str
    ptt_pwd: str
    board_name: str
    file_name: str
    max_retries: int = 3
    retry_delay: int = 5

class PTTCrawler:
    def __init__(self, config: CrawlerConfig):
        self.config = config
        self.ptt_bot = PTT.API()
        self.workbook: Optional[openpyxl.Workbook] = None
        self.sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None
        self.start_date = None
        self.start_index = None

        # 設置日誌
        self._setup_logging()

    def _setup_logging(self) -> None:
        """設置日誌系統"""
        log_file = Path('logs')
        log_file.mkdir(exist_ok=True)
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file / 'ptt_crawler.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def load_progress(self) -> None:
        """載入已完成的進度，找出最早的日期和 INDEX"""
        if os.path.exists(self.config.file_name):
            self.workbook = openpyxl.load_workbook(self.config.file_name)
            self.sheet = self.workbook.active
            dates = []

            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[2]:  # 日期和索引皆存在
                    dates.append((row[0], row[2]))

            if dates:
                self.start_date, self.start_index = dates[-1]
                self.start_date = datetime.strptime(self.start_date, '%Y-%m-%d').date()
                self.start_index -= 1  # 跳到該日期前一天和 (INDEX-1)
                self.logger.info(f"從 {self.start_date} (索引 {self.start_index}) 繼續爬取")
            else:
                self.start_date = None
                self.start_index = None
        else:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = '每日文章統計'
            self.sheet.append(['日期', '文章數量', '最早文章索引'])
            self.workbook.save(self.config.file_name)

    def save_daily_count(self, current_date: date, count: int, index: int) -> None:
        """儲存每日統計結果"""
        self.sheet.append([current_date.strftime('%Y-%m-%d'), count, index])
        self.workbook.save(self.config.file_name)
        self.logger.info(f"已完成 {current_date} 的文章抓取，共 {count} 篇 (最早索引 {index})")

    def retry_operation(self, operation, *args, **kwargs):
        """重試機制"""
        for attempt in range(self.config.max_retries):
            try:
                return operation(*args, **kwargs)
            except Exception as e:
                if attempt == self.config.max_retries - 1:
                    raise
                self.logger.warning(f"操作失敗，{self.config.retry_delay}秒後重試: {e}")
                time.sleep(self.config.retry_delay)

    def crawl(self) -> None:
        """執行爬蟲主程序"""
        try:
            self.load_progress()
            self.retry_operation(self.ptt_bot.login, self.config.ptt_id, self.config.ptt_pwd)
            self.logger.info("成功登入 PTT")

            latest_index = self.retry_operation(
                self.ptt_bot.get_newest_index,
                NewIndex.BOARD,
                self.config.board_name
            )
            self.logger.info(f"最新文章索引: {latest_index}")

            index = self.start_index or latest_index
            daily_count: Dict[date, int] = {}
            current_date = None
            current_index = index

            with tqdm(total=index, desc="爬取進度") as pbar:
                while current_index > 0:
                    try:
                        post = self.retry_operation(
                            self.ptt_bot.get_post,
                            self.config.board_name,
                            index=current_index
                        )
                        post_date_str = post.get('date')
                        post_date = datetime.strptime(post_date_str, '%a %b %d %H:%M:%S %Y').date()

                        if current_date is None:
                            current_date = post_date

                        if post_date != current_date:
                            # 儲存每日統計結果
                            self.save_daily_count(current_date, daily_count[current_date], current_index + 1)
                            current_date = post_date
                            daily_count[current_date] = 1
                        else:
                            daily_count[post_date] = daily_count.get(post_date, 0) + 1

                    except Exception as e:
                        self.logger.error(f"處理索引 {current_index} 時發生錯誤: {e}")

                    current_index -= 1
                    pbar.update(1)

            if current_date:
                self.save_daily_count(current_date, daily_count[current_date], current_index + 1)

            self.logger.info(f"所有資料已成功儲存至 '{self.config.file_name}'")

        except Exception as e:
            self.logger.error(f"爬蟲過程中發生錯誤: {e}")
            raise
        finally:
            try:
                self.ptt_bot.logout()
                self.logger.info("已登出 PTT")
            except Exception as e:
                self.logger.error(f"登出時發生錯誤: {e}")

# 使用範例
config = CrawlerConfig(
    ptt_id='IrohaKazama',
    ptt_pwd='eh586421973',
    board_name="Stock",
    file_name="ptt_data.xlsx"
)

crawler = PTTCrawler(config)
crawler.crawl()
